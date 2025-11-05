"""
Analytics Dashboard Views
Features: Percentile rankings, skill gap analysis, TAO rubric scoring
"""
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Avg, Count, Q, F, Max, Min, Sum, StdDev
from assessment.models import TestAttempt, Answer, TestCategory, ProctoringEvent, Test
from django.db.models.functions import TruncDate
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from collections import defaultdict, Counter
from django.utils import timezone
from datetime import timedelta
import json
import io
import base64

from assessment.models import (
    TestAttempt, Answer, TestCategory, ProctoringEvent, 
    Test, Question, QuestionTopic, Cohort, User
)

# For exports and visualizations
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter

# PDF generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.pdfgen import canvas

# Matplotlib for charts
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np

User = get_user_model()


MINIMUM_USERS_FOR_PERCENTILE = 5  # Minimum users needed for meaningful percentile comparison

@login_required
def user_analytics_dashboard(request):
    """
    User-facing analytics dashboard
    Shows: pass/fail, percentile ranking, skill gaps
    """
    user = request.user
    attempts = TestAttempt.objects.filter(
        user=user,
        status='completed'
    ).select_related('test__category').order_by('-completed_at')
    
    avg_score_decimal = attempts.aggregate(Avg('score'))['score__avg']
    context = {
        'attempts': attempts,
        'total_tests': attempts.count(),
        'passed_tests': attempts.filter(passed=True).count(),
        'avg_score': float(avg_score_decimal) if avg_score_decimal is not None else 0,
        'minimum_users_for_percentile': MINIMUM_USERS_FOR_PERCENTILE
    }
    
    # Calculate percentile ranking for each category
    category_percentiles = {}
    for category in TestCategory.objects.filter(is_active=True):
        percentile_data = calculate_user_percentile(user, category)
        category_percentiles[category.name] = percentile_data
    context['category_percentiles'] = category_percentiles
    
    # Skill gap analysis
    skill_gaps = analyze_skill_gaps(user)
    context['skill_gaps'] = skill_gaps
    
    # TAO rubric assessment
    tao_assessment = calculate_tao_rubric_score(user)
    context['tao_assessment'] = tao_assessment
    
    return render(request, 'assessment/analytics_dashboard.html', context)


@user_passes_test(lambda u: u.is_staff)
def admin_analytics_dashboard(request):
    """
    Comprehensive admin analytics dashboard with ALL statistics from Key_Statistics.docx
    """
    
    # Get filter parameters
    test_id = request.GET.get('test_id')
    cohort_id = request.GET.get('cohort_id')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Base queryset
    attempts_qs = TestAttempt.objects.filter(status='completed')
    
    # Apply filters
    if test_id:
        attempts_qs = attempts_qs.filter(test_id=test_id)
    if cohort_id:
        attempts_qs = attempts_qs.filter(cohort_id=cohort_id)
    if date_from:
        attempts_qs = attempts_qs.filter(completed_at__gte=date_from)
    if date_to:
        attempts_qs = attempts_qs.filter(completed_at__lte=date_to)
    
    # ===== 1. BASIC STATISTICS =====
    basic_stats = calculate_basic_statistics(attempts_qs)
    
    # ===== 2. SCORE STATISTICS =====
    score_stats = calculate_score_statistics(attempts_qs)
    
    # ===== 3. TIME STATISTICS =====
    time_stats = calculate_time_statistics(attempts_qs)
    
    # ===== 4. QUESTION STATISTICS =====
    question_stats = calculate_question_statistics(attempts_qs)
    
    # ===== 5. SECTION STATISTICS =====
    section_stats = calculate_section_statistics(attempts_qs)
    
    # ===== 6. COMPLETION STATISTICS =====
    completion_stats = calculate_completion_statistics(attempts_qs)
    
    # ===== 7. DEMOGRAPHIC STATISTICS =====
    demographic_stats = calculate_demographic_statistics(attempts_qs)
    
    # ===== 8. PLATFORM STATISTICS =====
    platform_stats = calculate_platform_statistics(attempts_qs)
    
    # ===== 9. RELIABILITY METRICS =====
    reliability_metrics = calculate_reliability_metrics(attempts_qs)
    
    # ===== 10. GENERATE VISUALIZATIONS =====
    charts = generate_dashboard_charts(
        basic_stats, score_stats, time_stats, question_stats, 
        section_stats, demographic_stats
    )
    
    # Get available filters
    available_tests = Test.objects.all()
    available_cohorts = Cohort.objects.all()
    
    context = {
        # Statistics
        'basic_stats': basic_stats,
        'score_stats': score_stats,
        'time_stats': time_stats,
        'question_stats': question_stats,
        'section_stats': section_stats,
        'completion_stats': completion_stats,
        'demographic_stats': demographic_stats,
        'platform_stats': platform_stats,
        'reliability_metrics': reliability_metrics,
        
        # Charts
        'charts': charts,
        
        # Filters
        'available_tests': available_tests,
        'available_cohorts': available_cohorts,
        'selected_test': test_id,
        'selected_cohort': cohort_id,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'admin/admin_analytics_enhanced.html', context)


def calculate_user_percentile(user, category):
    # Get user's average score in this category
    user_attempts = TestAttempt.objects.filter(
        user=user,
        test__category=category,
        status='completed',
        passed__isnull=False
    )
    
    if not user_attempts.exists():
        return {
            'percentile': None,
            'total_users': 0,
            'user_score': None,
            'has_sufficient_data': False,
            'message': 'No completed tests in this category'
        }
    
    user_avg_score = user_attempts.aggregate(Avg('score'))['score__avg']
    
    if user_avg_score is None:
        return {
            'percentile': None,
            'total_users': 0,
            'user_score': None,
            'has_sufficient_data': False,
            'message': 'Unable to calculate average score'
        }
    
    # Get all scores in this category from all users (excluding current user)
    all_scores = list(
        TestAttempt.objects.filter(
            test__category=category,
            status='completed',
            passed__isnull=False
        ).exclude(
            user=user
        ).values_list('score', flat=True)
    )
    
    total_users = len(all_scores) + 1  # Include current user
    
    # PRODUCTION READY: Check if we have enough users for meaningful comparison
    if len(all_scores) < MINIMUM_USERS_FOR_PERCENTILE - 1:
        return {
            'percentile': None,
            'total_users': total_users,
            'user_score': round(float(user_avg_score), 1),
            'has_sufficient_data': False,
            'message': f'Need at least {MINIMUM_USERS_FOR_PERCENTILE} users for percentile comparison'
        }
    
    # Add user's score to the list for calculation
    all_scores.append(float(user_avg_score))
    
    # Calculate percentile: percentage of scores below the user's score
    scores_below = sum(1 for score in all_scores if score < user_avg_score)
    percentile = (scores_below / len(all_scores)) * 100
    
    return {
        'percentile': round(percentile, 1),
        'total_users': total_users,
        'user_score': round(float(user_avg_score), 1),
        'has_sufficient_data': True,
        'message': None
    }


def analyze_skill_gaps(user):
    """
    Analyze user's performance by topic to identify skill gaps
    Returns topics where user scored < 60%
    
    NO CHANGES - This function works correctly as-is
    """
    skill_gaps = []
    
    # Get all user's completed attempts
    attempts = TestAttempt.objects.filter(
        user=user,
        status='completed'
    ).prefetch_related('answers__question__topic')
    
    # Aggregate performance by topic
    topic_performance = defaultdict(lambda: {'correct': 0, 'total': 0})
    
    for attempt in attempts:
        for answer in attempt.answers.all():
            topic = answer.question.topic
            topic_performance[topic]['total'] += 1
            if answer.is_correct:
                topic_performance[topic]['correct'] += 1
    
    # Identify gaps (< 60% accuracy)
    for topic, stats in topic_performance.items():
        if stats['total'] == 0:
            continue
        
        percentage = (stats['correct'] / stats['total']) * 100
        
        if percentage < 60:
            skill_gaps.append({
                'topic': topic.name,
                'category': topic.category.name,
                'avg_score': round(percentage, 1),
                'correct': stats['correct'],
                'total': stats['total'],
                'attempt_count': stats['total'],
                'questions_needed': max(1, int((0.6 * stats['total']) - stats['correct']))
            })
    
    # Sort by weakest first
    skill_gaps.sort(key=lambda x: x['avg_score'])
    
    return skill_gaps


def calculate_tao_rubric_score(user):
    """
    Calculate TAO-style rubric scores for the 4-stage assessment
    Returns readiness for each stage and overall certification readiness
    """
    rubric_scores = {
        'stage_1_cognitive': {'score': 0, 'max': 100, 'passed': False, 'percentile': None},
        'stage_2_detail': {'score': 0, 'max': 100, 'passed': False, 'percentile': None},
        'stage_3_trainability': {'score': 0, 'max': 100, 'passed': False, 'percentile': None},
        'stage_4_domain': {'score': 0, 'max': 100, 'passed': False, 'percentile': None},
        'overall_readiness': 0,
        'certification_ready': False
    }
    
    for stage_num in range(1, 5):
        category = TestCategory.objects.filter(stage_number=stage_num).first()
        if not category:
            continue
        
        # Get user's attempts for this stage
        attempts = TestAttempt.objects.filter(
            user=user,
            test__category=category,
            status='completed'
        )
        
        if attempts.exists():
            avg_score = attempts.aggregate(Avg('score'))['score__avg']
            passed = avg_score >= category.passing_score
            
            # UPDATED: Use new percentile calculation
            percentile_data = calculate_user_percentile(user, category)
            percentile_value = percentile_data['percentile'] if percentile_data['has_sufficient_data'] else None
            
            stage_key = f'stage_{stage_num}_{category.name.lower().split()[0]}'
            if stage_key in rubric_scores:
                rubric_scores[stage_key]['score'] = round(avg_score, 1)
                rubric_scores[stage_key]['passed'] = passed
                rubric_scores[stage_key]['percentile'] = percentile_value
    
    # Calculate overall readiness (average of all stages)
    stage_scores = [
        v['score'] for k, v in rubric_scores.items() 
        if k.startswith('stage_') and v['score'] > 0
    ]
    
    if stage_scores:
        overall = sum(stage_scores) / len(stage_scores)
        rubric_scores['overall_readiness'] = round(overall, 1)
        
        # Certification ready if all stages passed
        all_passed = all(
            v['passed'] for k, v in rubric_scores.items() 
            if k.startswith('stage_')
        )
        rubric_scores['certification_ready'] = all_passed
    
    return rubric_scores


def calculate_basic_statistics(attempts_qs):
    """Calculate basic overview statistics"""
    total_candidates = attempts_qs.values('user').distinct().count()
    total_attempts = attempts_qs.count()
    
    if total_attempts == 0:
        return {
            'total_candidates': 0,
            'total_attempts': 0,
            'pass_rate': 0,
            'failure_rate': 0,
        }
    
    passed_attempts = attempts_qs.filter(passed=True).count()
    pass_rate = (float(passed_attempts) / float(total_attempts) * 100) if total_attempts > 0 else 0
    failure_rate = 100 - pass_rate
    
    return {
        'total_candidates': total_candidates,
        'total_attempts': total_attempts,
        'passed_attempts': passed_attempts,
        'failed_attempts': total_attempts - passed_attempts,
        'pass_rate': round(pass_rate, 2),
        'failure_rate': round(failure_rate, 2),
    }


def calculate_score_statistics(attempts_qs):
    """Calculate all score-related statistics"""
    if not attempts_qs.exists():
        return {}
    
    scores = attempts_qs.values_list('score', flat=True)
    scores_list = [float(s) for s in scores if s is not None]
    
    if not scores_list:
        return {}
    
    # Calculate statistics
    avg_score = np.mean(scores_list)
    median_score = np.median(scores_list)
    std_dev = np.std(scores_list)
    highest_score = max(scores_list)
    lowest_score = min(scores_list)
    score_range = highest_score - lowest_score
    
    # Percentiles
    top_10_threshold = np.percentile(scores_list, 90)
    bottom_10_threshold = np.percentile(scores_list, 10)
    
    top_performers = len([s for s in scores_list if s >= top_10_threshold])
    weak_performers = len([s for s in scores_list if s <= bottom_10_threshold])
    
    # Score distribution
    score_distribution = {
        '0-20': len([s for s in scores_list if 0 <= s < 20]),
        '20-40': len([s for s in scores_list if 20 <= s < 40]),
        '40-60': len([s for s in scores_list if 40 <= s < 60]),
        '60-80': len([s for s in scores_list if 60 <= s < 80]),
        '80-100': len([s for s in scores_list if 80 <= s <= 100]),
    }
    
    return {
        'average_score': round(avg_score, 2),
        'median_score': round(median_score, 2),
        'score_std_dev': round(std_dev, 2),
        'highest_score': round(highest_score, 2),
        'lowest_score': round(lowest_score, 2),
        'score_range': round(score_range, 2),
        'top_10_threshold': round(top_10_threshold, 2),
        'bottom_10_threshold': round(bottom_10_threshold, 2),
        'top_performers_count': top_performers,
        'weak_performers_count': weak_performers,
        'score_distribution': score_distribution,
    }


def calculate_time_statistics(attempts_qs):
    """Calculate time-related statistics"""
    if not attempts_qs.exists():
        return {}
    
    attempts_with_time = attempts_qs.exclude(time_spent_seconds__isnull=True)
    
    if not attempts_with_time.exists():
        return {}
    
    time_seconds = attempts_with_time.values_list('time_spent_seconds', flat=True)
    
    time_minutes = [float(t) / 60 for t in time_seconds if t]
    
    if not time_minutes:
        return {}
    
    avg_time = np.mean(time_minutes)
    median_time = np.median(time_minutes)
    min_time = min(time_minutes)
    max_time = max(time_minutes)
    
    # Time distribution (fast, average, slow)
    fast_threshold = np.percentile(time_minutes, 33)
    slow_threshold = np.percentile(time_minutes, 67)
    
    fast_finishers = len([t for t in time_minutes if t <= fast_threshold])
    average_finishers = len([t for t in time_minutes if fast_threshold < t <= slow_threshold])
    slow_finishers = len([t for t in time_minutes if t > slow_threshold])
    
    return {
        'average_time_minutes': round(avg_time, 2),
        'median_time_minutes': round(median_time, 2),
        'min_time_minutes': round(min_time, 2),
        'max_time_minutes': round(max_time, 2),
        'time_distribution': {
            'fast': fast_finishers,
            'average': average_finishers,
            'slow': slow_finishers,
        }
    }


def calculate_question_statistics(attempts_qs):
    """Calculate question-level statistics"""
    if not attempts_qs.exists():
        return {}
    
    # Get all answers for completed attempts
    answers = Answer.objects.filter(
        attempt__in=attempts_qs,
        question__isnull=False
    ).select_related('question')
    
    if not answers.exists():
        return {}
    
    # Question-wise correct response rate
    question_stats = {}
    for answer in answers:
        q_id = answer.question.id
        if q_id not in question_stats:
            question_stats[q_id] = {
                'question': answer.question,
                'total': 0,
                'correct': 0,
            }
        question_stats[q_id]['total'] += 1
        if answer.is_correct:
            question_stats[q_id]['correct'] += 1
    
    # Calculate percentages and difficulty
    question_performance = []
    for q_id, stats in question_stats.items():
        correct_rate = (stats['correct'] / stats['total'] * 100) if stats['total'] > 0 else 0
        question_performance.append({
            'question_id': q_id,
            'question_text': stats['question'].question_text[:100],
            'correct_rate': round(correct_rate, 2),
            'total_attempts': stats['total'],
            'difficulty_index': round(100 - correct_rate, 2),  # Higher = more difficult
        })
    
    # Sort to find hardest and easiest questions
    question_performance.sort(key=lambda x: x['correct_rate'])
    
    most_missed = question_performance[:5]  # 5 hardest
    most_correctly_answered = question_performance[-5:]  # 5 easiest
    most_correctly_answered.reverse()
    
    # Item Discrimination Index (simplified)
    # Top 27% vs Bottom 27% comparison
    all_scores = list(attempts_qs.values_list('score', flat=True))
    if len(all_scores) >= 10:
        top_27_threshold = np.percentile(all_scores, 73)
        bottom_27_threshold = np.percentile(all_scores, 27)
        
        discrimination_indices = calculate_discrimination_indices(
            attempts_qs, top_27_threshold, bottom_27_threshold
        )
    else:
        discrimination_indices = []
    
    return {
        'question_performance': question_performance,
        'most_missed_questions': most_missed,
        'most_correctly_answered': most_correctly_answered,
        'discrimination_indices': discrimination_indices,
    }


def calculate_discrimination_indices(attempts_qs, top_threshold, bottom_threshold):
    """Calculate item discrimination index for questions"""
    top_group = attempts_qs.filter(score__gte=top_threshold)
    bottom_group = attempts_qs.filter(score__lte=bottom_threshold)
    
    discrimination_data = []
    
    # Get all questions
    questions = Question.objects.all()[:20]  # Limit for performance
    
    for question in questions:
        top_correct = Answer.objects.filter(
            attempt__in=top_group,
            question=question,
            is_correct=True
        ).count()

        bottom_correct = Answer.objects.filter(
            attempt__in=bottom_group,
            question=question,
            is_correct=True
        ).count()

        top_total = Answer.objects.filter(
            attempt__in=top_group,
            question=question
        ).count()

        bottom_total = Answer.objects.filter(
            attempt__in=bottom_group,
            question=question
        ).count()
        
        if top_total > 0 and bottom_total > 0:
            top_prop = top_correct / top_total
            bottom_prop = bottom_correct / bottom_total
            discrimination = top_prop - bottom_prop
            
            discrimination_data.append({
                'question_text': question.question_text[:80],
                'discrimination_index': round(discrimination, 3),
                'interpretation': interpret_discrimination(discrimination)
            })
    
    return sorted(discrimination_data, key=lambda x: x['discrimination_index'], reverse=True)


def interpret_discrimination(index):
    """Interpret discrimination index value"""
    if index >= 0.40:
        return "Excellent"
    elif index >= 0.30:
        return "Good"
    elif index >= 0.20:
        return "Acceptable"
    elif index >= 0.10:
        return "Marginal"
    else:
        return "Poor - needs revision"


def calculate_section_statistics(attempts_qs):
    """Calculate section/category-wise statistics"""
    if not attempts_qs.exists():
        return {}
    
    sections = {}
    
    for category in TestCategory.objects.all():
        category_attempts = attempts_qs.filter(test__category=category)
        
        if category_attempts.exists():
            avg_score_decimal = category_attempts.aggregate(Avg('score'))['score__avg']
            avg_score = float(avg_score_decimal) if avg_score_decimal else 0
            
            pass_rate = category_attempts.filter(passed=True).count() / category_attempts.count() * 100
            
            # Calculate difficulty index (100 - average score)
            difficulty_index = 100 - float(avg_score or 0)
            
            sections[category.name] = {
                'average_score': round(float(avg_score or 0), 2),
                'pass_rate': round(pass_rate, 2),
                'attempts_count': category_attempts.count(),
                'difficulty_index': round(difficulty_index, 2),
                'interpretation': interpret_difficulty(difficulty_index)
            }
    
    return sections


def interpret_difficulty(index):
    """Interpret difficulty index"""
    if index < 20:
        return "Very Easy"
    elif index < 40:
        return "Easy"
    elif index < 60:
        return "Moderate"
    elif index < 80:
        return "Difficult"
    else:
        return "Very Difficult"


def calculate_completion_statistics(attempts_qs):
    """Calculate completion and drop-off statistics"""
    total_started = TestAttempt.objects.count()
    completed = attempts_qs.count()
    incomplete = TestAttempt.objects.filter(status='in_progress').count()
    abandoned = TestAttempt.objects.filter(status='abandoned').count()
    
    if total_started == 0:
        return {}
    
    completion_rate = (completed / total_started * 100) if total_started > 0 else 0
    drop_off_rate = 100 - completion_rate
    
    # Retake statistics
    users_with_retakes = TestAttempt.objects.values('user', 'test').annotate(
        attempt_count=Count('id')
    ).filter(attempt_count__gt=1)
    
    retake_data = []
    for record in users_with_retakes[:20]:  # Limit for performance
        user_attempts = TestAttempt.objects.filter(
            user_id=record['user'],
            test_id=record['test'],
            status='completed'
        ).order_by('completed_at')
        
        if user_attempts.count() >= 2:
            first_score = float(user_attempts.first().score or 0)
            last_score = float(user_attempts.last().score or 0)
            improvement = last_score - first_score
            
            retake_data.append({
                'user_id': record['user'],
                'attempts': record['attempt_count'],
                'first_score': round(first_score, 2),
                'last_score': round(last_score, 2),
                'improvement': round(improvement, 2)
            })
    
    # Skipped questions
    total_questions_seen = Answer.objects.filter(attempt__in=attempts_qs).count()
    skipped_questions = Answer.objects.filter(
        attempt__in=attempts_qs,
        selected_answer__isnull=True
    ).count()
    
    skipped_rate = (skipped_questions / total_questions_seen * 100) if total_questions_seen > 0 else 0
    
    return {
        'total_started': total_started,
        'completed': completed,
        'incomplete': incomplete,
        'abandoned': abandoned,
        'completion_rate': round(completion_rate, 2),
        'drop_off_rate': round(drop_off_rate, 2),
        'retake_data': retake_data,
        'skipped_questions_rate': round(skipped_rate, 2),
    }


def calculate_demographic_statistics(attempts_qs):
    """Calculate demographic statistics"""
    if not attempts_qs.exists():
        return {}
    
    # Age distribution (if available)
    users = User.objects.filter(test_attempts__in=attempts_qs).distinct()
    
    age_data = []
    for user in users:
        if hasattr(user, 'profile') and hasattr(user.profile, 'date_of_birth'):
            if user.profile.date_of_birth:
                age = (timezone.now().date() - user.profile.date_of_birth).days // 365
                age_data.append(age)
    
    age_stats = {}
    if age_data:
        age_stats = {
            'average_age': round(np.mean(age_data), 1),
            'age_distribution': {
                '18-25': len([a for a in age_data if 18 <= a < 25]),
                '26-35': len([a for a in age_data if 26 <= a < 35]),
                '36-45': len([a for a in age_data if 36 <= a < 45]),
                '46+': len([a for a in age_data if a >= 46]),
            }
        }
    
    # Gender distribution (if available)
    gender_dist = {}
    if hasattr(User, 'profile'):
        for gender_choice in ['M', 'F', 'O']:
            count = users.filter(profile__gender=gender_choice).count()
            if count > 0:
                gender_dist[gender_choice] = count
    
    # Education level (if available)
    education_dist = {}
    if hasattr(User, 'profile'):
        education_levels = users.values('profile__education_level').annotate(
            count=Count('id')
        )
        for level in education_levels:
            if level['profile__education_level']:
                education_dist[level['profile__education_level']] = level['count']
    
    return {
        'age_stats': age_stats,
        'gender_distribution': gender_dist,
        'education_distribution': education_dist,
    }


def calculate_platform_statistics(attempts_qs):
    """Calculate platform/device statistics"""
    if not attempts_qs.exists():
        return {}
    
    # Device/Platform distribution
    platform_data = attempts_qs.values('user_agent').annotate(
        count=Count('id')
    ).order_by('-count')
    
    device_types = {
        'Desktop': 0,
        'Mobile': 0,
        'Tablet': 0,
        'Unknown': 0
    }
    
    for platform in platform_data:
        user_agent = platform['user_agent'] or ''
        count = platform['count']
        
        if 'Mobile' in user_agent or 'Android' in user_agent:
            device_types['Mobile'] += count
        elif 'iPad' in user_agent or 'Tablet' in user_agent:
            device_types['Tablet'] += count
        elif user_agent:
            device_types['Desktop'] += count
        else:
            device_types['Unknown'] += count
    
    return {
        'device_distribution': device_types,
    }


def calculate_reliability_metrics(attempts_qs):
    """Calculate test reliability metrics (Cronbach's Alpha approximation)"""
    if not attempts_qs.exists() or attempts_qs.count() < 10:
        return {}
    
    # Get variance of total scores
    scores = [float(s) for s in attempts_qs.values_list('score', flat=True) if s]
    
    if len(scores) < 10:
        return {}
    
    total_variance = np.var(scores)
    
    # Get variance of individual items (questions)
    # Simplified calculation
    questions = Question.objects.all()[:20]  # Sample
    item_variances = []
    
    for question in questions:
        answers = Answer.objects.filter(
            attempt__in=attempts_qs,
            question=question
        )
        
        if answers.count() > 5:
            correct_list = [1 if a.is_correct else 0 for a in answers]
            if len(correct_list) > 0:
                item_variances.append(np.var(correct_list))
    
    if item_variances and total_variance > 0:
        k = len(item_variances)
        sum_item_var = sum(item_variances)
        
        # Cronbach's Alpha formula
        alpha = (k / (k - 1)) * (1 - (sum_item_var / total_variance))
        
        reliability_interpretation = interpret_reliability(alpha)
    else:
        alpha = None
        reliability_interpretation = "Insufficient data"
    
    return {
        'cronbach_alpha': round(alpha, 3) if alpha else None,
        'interpretation': reliability_interpretation,
    }


def interpret_reliability(alpha):
    """Interpret Cronbach's Alpha value"""
    if alpha is None:
        return "Cannot calculate"
    elif alpha >= 0.9:
        return "Excellent reliability"
    elif alpha >= 0.8:
        return "Good reliability"
    elif alpha >= 0.7:
        return "Acceptable reliability"
    elif alpha >= 0.6:
        return "Questionable reliability"
    else:
        return "Poor reliability - test needs revision"


def generate_dashboard_charts(basic_stats, score_stats, time_stats, 
                              question_stats, section_stats, demographic_stats):
    """Generate all visualization charts as base64 images"""
    charts = {}
    
    # Set style
    sns.set_style("whitegrid")
    
    # 1. Pass/Fail Pie Chart
    if basic_stats and basic_stats.get('total_attempts', 0) > 0:
        charts['pass_fail_pie'] = create_pass_fail_pie(basic_stats)
    
    # 2. Score Distribution Histogram
    if score_stats and score_stats.get('score_distribution'):
        charts['score_distribution'] = create_score_distribution(score_stats)
    
    # 3. Time Distribution
    if time_stats and time_stats.get('time_distribution'):
        charts['time_distribution'] = create_time_distribution(time_stats)
    
    # 4. Section Performance Bar Chart
    if section_stats:
        charts['section_performance'] = create_section_performance(section_stats)
    
    # 5. Question Difficulty Heatmap (top 20 hardest)
    if question_stats and question_stats.get('question_performance'):
        charts['question_difficulty'] = create_question_difficulty(question_stats)
    
    # 6. Demographics
    if demographic_stats:
        if demographic_stats.get('age_stats', {}).get('age_distribution'):
            charts['age_distribution'] = create_age_distribution(demographic_stats)
        if demographic_stats.get('gender_distribution'):
            charts['gender_distribution'] = create_gender_distribution(demographic_stats)
    
    return charts


def create_pass_fail_pie(basic_stats):
    """Create pass/fail pie chart"""
    fig, ax = plt.subplots(figsize=(8, 6))
    
    sizes = [basic_stats['passed_attempts'], basic_stats['failed_attempts']]
    labels = ['Passed', 'Failed']
    colors = ['#4CAF50', '#f44336']
    explode = (0.1, 0)
    
    ax.pie(sizes, explode=explode, labels=labels, colors=colors, autopct='%1.1f%%',
           shadow=True, startangle=90)
    ax.axis('equal')
    plt.title('Pass/Fail Distribution', fontsize=16, fontweight='bold')
    
    return fig_to_base64(fig)


def create_score_distribution(score_stats):
    """Create score distribution bar chart"""
    fig, ax = plt.subplots(figsize=(10, 6))
    
    distribution = score_stats['score_distribution']
    ranges = list(distribution.keys())
    counts = list(distribution.values())
    
    colors_list = ['#f44336', '#FF9800', '#FFC107', '#8BC34A', '#4CAF50']
    bars = ax.bar(ranges, counts, color=colors_list, edgecolor='black', linewidth=1.2)
    
    ax.set_xlabel('Score Range (%)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Number of Candidates', fontsize=12, fontweight='bold')
    ax.set_title('Score Distribution', fontsize=16, fontweight='bold')
    ax.grid(axis='y', alpha=0.3)
    
    # Add value labels on bars
    for bar in bars:
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2., height,
                f'{int(height)}',
                ha='center', va='bottom', fontweight='bold')
    
    plt.tight_layout()
    return fig_to_base64(fig)


def create_time_distribution(time_stats):
    """Create time distribution pie chart"""
    fig, ax = plt.subplots(figsize=(8, 6))
    
    distribution = time_stats['time_distribution']
    sizes = [distribution['fast'], distribution['average'], distribution['slow']]
    labels = ['Fast Finishers', 'Average', 'Slow Finishers']
    colors = ['#2196F3', '#FFC107', '#FF5722']
    
    ax.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', 
           shadow=True, startangle=90)
    ax.axis('equal')
    plt.title('Completion Time Distribution', fontsize=16, fontweight='bold')
    
    return fig_to_base64(fig)


def create_section_performance(section_stats):
    """Create section performance bar chart"""
    fig, ax = plt.subplots(figsize=(12, 6))
    
    sections = list(section_stats.keys())
    scores = [section_stats[s]['average_score'] for s in sections]
    
    bars = ax.barh(sections, scores, color='#667eea', edgecolor='black', linewidth=1.2)
    
    ax.set_xlabel('Average Score (%)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Test Section', fontsize=12, fontweight='bold')
    ax.set_title('Section-wise Average Performance', fontsize=16, fontweight='bold')
    ax.set_xlim(0, 100)
    ax.grid(axis='x', alpha=0.3)
    
    # Add value labels
    for i, bar in enumerate(bars):
        width = bar.get_width()
        ax.text(width + 2, bar.get_y() + bar.get_height()/2.,
                f'{scores[i]:.1f}%',
                ha='left', va='center', fontweight='bold')
    
    plt.tight_layout()
    return fig_to_base64(fig)


def create_question_difficulty(question_stats):
    """Create question difficulty chart (hardest 10 questions)"""
    fig, ax = plt.subplots(figsize=(12, 8))
    
    questions = question_stats['most_missed_questions'][:10]
    
    if not questions:
        return None
    
    q_labels = [f"Q{i+1}" for i in range(len(questions))]
    difficulty = [q['difficulty_index'] for q in questions]
    
    colors_list = ['#f44336' if d > 70 else '#FF9800' if d > 50 else '#FFC107' 
                   for d in difficulty]
    
    bars = ax.barh(q_labels, difficulty, color=colors_list, edgecolor='black', linewidth=1.2)
    
    ax.set_xlabel('Difficulty Index (Higher = More Difficult)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Question', fontsize=12, fontweight='bold')
    ax.set_title('Top 10 Most Difficult Questions', fontsize=16, fontweight='bold')
    ax.set_xlim(0, 100)
    ax.grid(axis='x', alpha=0.3)
    
    # Add value labels
    for i, bar in enumerate(bars):
        width = bar.get_width()
        ax.text(width + 2, bar.get_y() + bar.get_height()/2.,
                f'{difficulty[i]:.1f}',
                ha='left', va='center', fontweight='bold')
    
    plt.tight_layout()
    return fig_to_base64(fig)


def create_age_distribution(demographic_stats):
    """Create age distribution bar chart"""
    fig, ax = plt.subplots(figsize=(10, 6))
    
    age_dist = demographic_stats['age_stats']['age_distribution']
    ranges = list(age_dist.keys())
    counts = list(age_dist.values())
    
    bars = ax.bar(ranges, counts, color='#9C27B0', edgecolor='black', linewidth=1.2)
    
    ax.set_xlabel('Age Range', fontsize=12, fontweight='bold')
    ax.set_ylabel('Number of Candidates', fontsize=12, fontweight='bold')
    ax.set_title('Age Distribution', fontsize=16, fontweight='bold')
    ax.grid(axis='y', alpha=0.3)
    
    for bar in bars:
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2., height,
                f'{int(height)}',
                ha='center', va='bottom', fontweight='bold')
    
    plt.tight_layout()
    return fig_to_base64(fig)


def create_gender_distribution(demographic_stats):
    """Create gender distribution pie chart"""
    fig, ax = plt.subplots(figsize=(8, 6))
    
    gender_dist = demographic_stats['gender_distribution']
    
    labels_map = {'M': 'Male', 'F': 'Female', 'O': 'Other'}
    labels = [labels_map.get(k, k) for k in gender_dist.keys()]
    sizes = list(gender_dist.values())
    colors = ['#2196F3', '#E91E63', '#9C27B0']
    
    ax.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%',
           shadow=True, startangle=90)
    ax.axis('equal')
    plt.title('Gender Distribution', fontsize=16, fontweight='bold')
    
    return fig_to_base64(fig)


def fig_to_base64(fig):
    """Convert matplotlib figure to base64 string"""
    buffer = io.BytesIO()
    fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
    buffer.seek(0)
    image_base64 = base64.b64encode(buffer.read()).decode()
    plt.close(fig)
    return f"data:image/png;base64,{image_base64}"


# ==================== EXPORT FUNCTIONS ====================

@user_passes_test(lambda u: u.is_staff)
def export_analytics_excel(request):
    """
    Export comprehensive analytics to Excel with multiple sheets and charts
    """
    # Get same filters as dashboard
    test_id = request.GET.get('test_id')
    cohort_id = request.GET.get('cohort_id')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    attempts_qs = TestAttempt.objects.filter(status='completed')
    
    if test_id:
        attempts_qs = attempts_qs.filter(test_id=test_id)
    if cohort_id:
        attempts_qs = attempts_qs.filter(cohort_id=cohort_id)
    if date_from:
        attempts_qs = attempts_qs.filter(completed_at__gte=date_from)
    if date_to:
        attempts_qs = attempts_qs.filter(completed_at__lte=date_to)
    
    # Calculate all statistics
    basic_stats = calculate_basic_statistics(attempts_qs)
    score_stats = calculate_score_statistics(attempts_qs)
    time_stats = calculate_time_statistics(attempts_qs)
    question_stats = calculate_question_statistics(attempts_qs)
    section_stats = calculate_section_statistics(attempts_qs)
    completion_stats = calculate_completion_statistics(attempts_qs)
    demographic_stats = calculate_demographic_statistics(attempts_qs)
    platform_stats = calculate_platform_statistics(attempts_qs)
    reliability_metrics = calculate_reliability_metrics(attempts_qs)
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Summary Dashboard
    ws_summary = wb.active
    ws_summary.title = "Summary"
    create_summary_sheet(ws_summary, basic_stats, score_stats, time_stats, 
                        header_fill, header_font, title_font)
    
    # Sheet 2: All Test Attempts
    ws_attempts = wb.create_sheet("All Attempts")
    create_attempts_sheet(ws_attempts, attempts_qs, header_fill, header_font)
    
    # Sheet 3: Score Analysis
    ws_scores = wb.create_sheet("Score Analysis")
    create_score_analysis_sheet(ws_scores, score_stats, header_fill, header_font, title_font)
    
    # Sheet 4: Question Performance
    ws_questions = wb.create_sheet("Question Performance")
    create_question_performance_sheet(ws_questions, question_stats, header_fill, header_font)
    
    # Sheet 5: Section Analysis
    ws_sections = wb.create_sheet("Section Analysis")
    create_section_analysis_sheet(ws_sections, section_stats, header_fill, header_font)
    
    # Sheet 6: Demographics
    ws_demo = wb.create_sheet("Demographics")
    create_demographics_sheet(ws_demo, demographic_stats, header_fill, header_font)
    
    # Sheet 7: Time Analysis
    ws_time = wb.create_sheet("Time Analysis")
    create_time_analysis_sheet(ws_time, time_stats, header_fill, header_font)
    
    # Sheet 8: Reliability Metrics
    ws_reliability = wb.create_sheet("Reliability")
    create_reliability_sheet(ws_reliability, reliability_metrics, header_fill, header_font)
    
    # Save to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'analytics_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


def create_summary_sheet(ws, basic_stats, score_stats, time_stats, 
                        header_fill, header_font, title_font):
    """Create summary sheet in Excel"""
    ws['A1'] = 'COMPREHENSIVE ANALYTICS SUMMARY'
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    
    row = 3
    
    # Basic Statistics
    ws[f'A{row}'] = 'BASIC STATISTICS'
    ws[f'A{row}'].font = title_font
    row += 1
    
    ws[f'A{row}'] = 'Metric'
    ws[f'B{row}'] = 'Value'
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].font = header_font
    row += 1
    
    if basic_stats:
        for key, value in basic_stats.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = value
            row += 1
    
    row += 2
    
    # Score Statistics
    ws[f'A{row}'] = 'SCORE STATISTICS'
    ws[f'A{row}'].font = title_font
    row += 1
    
    ws[f'A{row}'] = 'Metric'
    ws[f'B{row}'] = 'Value'
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].font = header_font
    row += 1
    
    if score_stats:
        for key, value in score_stats.items():
            if key != 'score_distribution':
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = value
                row += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20


def create_attempts_sheet(ws, attempts_qs, header_fill, header_font):
    """Create detailed attempts sheet"""
    headers = ['User', 'Test', 'Category', 'Score (%)', 'Passed', 'Time (min)', 
               'Completed At', 'Flagged', 'IP Address']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for attempt in attempts_qs.select_related('user', 'test__category'):
        ws.append([
            attempt.user.username,
            attempt.test.title,
            attempt.test.category.name if attempt.test.category else 'N/A',
            round(float(attempt.score or 0), 2),
            'Yes' if attempt.passed else 'No',
            round((attempt.time_spent_seconds or 0) / 60, 2),
            attempt.completed_at.strftime('%Y-%m-%d %H:%M') if attempt.completed_at else 'N/A',
            'Yes' if attempt.flagged_for_plagiarism else 'No',
            attempt.ip_address or 'N/A'
        ])
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_score_analysis_sheet(ws, score_stats, header_fill, header_font, title_font):
    """Create score analysis sheet with distribution"""
    ws['A1'] = 'SCORE ANALYSIS'
    ws['A1'].font = title_font
    
    row = 3
    ws[f'A{row}'] = 'Metric'
    ws[f'B{row}'] = 'Value'
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].font = header_font
    row += 1
    
    if score_stats:
        for key, value in score_stats.items():
            if key == 'score_distribution':
                row += 2
                ws[f'A{row}'] = 'SCORE DISTRIBUTION'
                ws[f'A{row}'].font = title_font
                row += 1
                ws[f'A{row}'] = 'Range'
                ws[f'B{row}'] = 'Count'
                ws[f'A{row}'].fill = header_fill
                ws[f'B{row}'].fill = header_fill
                ws[f'A{row}'].font = header_font
                ws[f'B{row}'].font = header_font
                row += 1
                for range_key, count in value.items():
                    ws[f'A{row}'] = range_key
                    ws[f'B{row}'] = count
                    row += 1
            else:
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = value
                row += 1
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20


def create_question_performance_sheet(ws, question_stats, header_fill, header_font):
    """Create question performance sheet"""
    headers = ['Question ID', 'Question Text', 'Correct Rate (%)', 
               'Total Attempts', 'Difficulty Index']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    if question_stats and question_stats.get('question_performance'):
        for q in question_stats['question_performance']:
            ws.append([
                q['question_id'],
                q['question_text'],
                q['correct_rate'],
                q['total_attempts'],
                q['difficulty_index']
            ])
    
    # Auto-adjust columns
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20


def create_section_analysis_sheet(ws, section_stats, header_fill, header_font):
    """Create section analysis sheet"""
    headers = ['Section', 'Average Score (%)', 'Pass Rate (%)', 
               'Attempts', 'Difficulty Index', 'Interpretation']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for section_name, stats in section_stats.items():
        ws.append([
            section_name,
            stats['average_score'],
            stats['pass_rate'],
            stats['attempts_count'],
            stats['difficulty_index'],
            stats['interpretation']
        ])
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25


def create_demographics_sheet(ws, demographic_stats, header_fill, header_font):
    """Create demographics sheet"""
    row = 1
    ws[f'A{row}'] = 'DEMOGRAPHIC ANALYSIS'
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 2
    
    # Age distribution
    if demographic_stats.get('age_stats'):
        ws[f'A{row}'] = 'AGE DISTRIBUTION'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Age Range'
        ws[f'B{row}'] = 'Count'
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        row += 1
        
        age_dist = demographic_stats['age_stats'].get('age_distribution', {})
        for age_range, count in age_dist.items():
            ws[f'A{row}'] = age_range
            ws[f'B{row}'] = count
            row += 1
        
        row += 2
    
    # Gender distribution
    if demographic_stats.get('gender_distribution'):
        ws[f'A{row}'] = 'GENDER DISTRIBUTION'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Gender'
        ws[f'B{row}'] = 'Count'
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        row += 1
        
        for gender, count in demographic_stats['gender_distribution'].items():
            gender_map = {'M': 'Male', 'F': 'Female', 'O': 'Other'}
            ws[f'A{row}'] = gender_map.get(gender, gender)
            ws[f'B{row}'] = count
            row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15


def create_time_analysis_sheet(ws, time_stats, header_fill, header_font):
    """Create time analysis sheet"""
    ws['A1'] = 'TIME ANALYSIS'
    ws['A1'].font = Font(bold=True, size=14)
    
    row = 3
    ws[f'A{row}'] = 'Metric'
    ws[f'B{row}'] = 'Value (minutes)'
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].font = header_font
    row += 1
    
    if time_stats:
        ws[f'A{row}'] = 'Average Time'
        ws[f'B{row}'] = time_stats.get('average_time_minutes', 0)
        row += 1
        
        ws[f'A{row}'] = 'Median Time'
        ws[f'B{row}'] = time_stats.get('median_time_minutes', 0)
        row += 1
        
        ws[f'A{row}'] = 'Minimum Time'
        ws[f'B{row}'] = time_stats.get('min_time_minutes', 0)
        row += 1
        
        ws[f'A{row}'] = 'Maximum Time'
        ws[f'B{row}'] = time_stats.get('max_time_minutes', 0)
        row += 1
        
        if time_stats.get('time_distribution'):
            row += 2
            ws[f'A{row}'] = 'TIME DISTRIBUTION'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            ws[f'A{row}'] = 'Category'
            ws[f'B{row}'] = 'Count'
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'].fill = header_fill
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'].font = header_font
            row += 1
            
            for category, count in time_stats['time_distribution'].items():
                ws[f'A{row}'] = category.title()
                ws[f'B{row}'] = count
                row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20


def create_reliability_sheet(ws, reliability_metrics, header_fill, header_font):
    """Create reliability metrics sheet"""
    ws['A1'] = 'RELIABILITY METRICS'
    ws['A1'].font = Font(bold=True, size=14)
    
    row = 3
    ws[f'A{row}'] = 'Metric'
    ws[f'B{row}'] = 'Value'
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].font = header_font
    row += 1
    
    if reliability_metrics:
        ws[f'A{row}'] = "Cronbach's Alpha"
        ws[f'B{row}'] = reliability_metrics.get('cronbach_alpha', 'N/A')
        row += 1
        
        ws[f'A{row}'] = 'Interpretation'
        ws[f'B{row}'] = reliability_metrics.get('interpretation', 'N/A')
        row += 1
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30


@user_passes_test(lambda u: u.is_staff)
def export_analytics_pdf(request):
    """
    Export comprehensive analytics to PDF report
    """
    # Get filters
    test_id = request.GET.get('test_id')
    cohort_id = request.GET.get('cohort_id')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    attempts_qs = TestAttempt.objects.filter(status='completed')
    
    if test_id:
        attempts_qs = attempts_qs.filter(test_id=test_id)
    if cohort_id:
        attempts_qs = attempts_qs.filter(cohort_id=cohort_id)
    if date_from:
        attempts_qs = attempts_qs.filter(completed_at__gte=date_from)
    if date_to:
        attempts_qs = attempts_qs.filter(completed_at__lte=date_to)
    
    # Calculate statistics
    basic_stats = calculate_basic_statistics(attempts_qs)
    score_stats = calculate_score_statistics(attempts_qs)
    time_stats = calculate_time_statistics(attempts_qs)
    question_stats = calculate_question_statistics(attempts_qs)
    section_stats = calculate_section_statistics(attempts_qs)
    completion_stats = calculate_completion_statistics(attempts_qs)
    demographic_stats = calculate_demographic_statistics(attempts_qs)
    reliability_metrics = calculate_reliability_metrics(attempts_qs)
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.75*inch, bottomMargin=0.75*inch)
    
    # Container for PDF elements
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#4472C4'),
        spaceAfter=30,
        alignment=1  # Center
    )
    elements.append(Paragraph('COMPREHENSIVE ANALYTICS REPORT', title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Report metadata
    meta_data = [
        ['Report Generated:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['Total Attempts Analyzed:', str(basic_stats.get('total_attempts', 0))],
    ]
    meta_table = Table(meta_data, colWidths=[2*inch, 4*inch])
    meta_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
    ]))
    elements.append(meta_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # 1. Basic Statistics Section
    elements.append(Paragraph('1. BASIC STATISTICS', styles['Heading2']))
    if basic_stats:
        basic_data = [
            ['Metric', 'Value'],
            ['Total Candidates', str(basic_stats.get('total_candidates', 0))],
            ['Total Attempts', str(basic_stats.get('total_attempts', 0))],
            ['Passed Attempts', str(basic_stats.get('passed_attempts', 0))],
            ['Failed Attempts', str(basic_stats.get('failed_attempts', 0))],
            ['Pass Rate', f"{basic_stats.get('pass_rate', 0)}%"],
            ['Failure Rate', f"{basic_stats.get('failure_rate', 0)}%"],
        ]
        basic_table = create_pdf_table(basic_data)
        elements.append(basic_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # 2. Score Statistics Section
    elements.append(Paragraph('2. SCORE STATISTICS', styles['Heading2']))
    if score_stats:
        score_data = [
            ['Metric', 'Value'],
            ['Average Score', f"{score_stats.get('average_score', 0)}%"],
            ['Median Score', f"{score_stats.get('median_score', 0)}%"],
            ['Standard Deviation', f"{score_stats.get('score_std_dev', 0)}"],
            ['Highest Score', f"{score_stats.get('highest_score', 0)}%"],
            ['Lowest Score', f"{score_stats.get('lowest_score', 0)}%"],
            ['Score Range', f"{score_stats.get('score_range', 0)}"],
            ['Top Performers (Top 10%)', str(score_stats.get('top_performers_count', 0))],
            ['Weak Performers (Bottom 10%)', str(score_stats.get('weak_performers_count', 0))],
        ]
        score_table = create_pdf_table(score_data)
        elements.append(score_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # 3. Time Statistics
    elements.append(Paragraph('3. TIME STATISTICS', styles['Heading2']))
    if time_stats:
        time_data = [
            ['Metric', 'Value (minutes)'],
            ['Average Time', f"{time_stats.get('average_time_minutes', 0)}"],
            ['Median Time', f"{time_stats.get('median_time_minutes', 0)}"],
            ['Minimum Time', f"{time_stats.get('min_time_minutes', 0)}"],
            ['Maximum Time', f"{time_stats.get('max_time_minutes', 0)}"],
        ]
        time_table = create_pdf_table(time_data)
        elements.append(time_table)
    elements.append(PageBreak())
    
    # 4. Section Performance
    elements.append(Paragraph('4. SECTION PERFORMANCE', styles['Heading2']))
    if section_stats:
        section_data = [['Section', 'Avg Score', 'Pass Rate', 'Difficulty']]
        for section_name, stats in section_stats.items():
            section_data.append([
                section_name[:30],
                f"{stats['average_score']}%",
                f"{stats['pass_rate']}%",
                stats['interpretation']
            ])
        section_table = create_pdf_table(section_data)
        elements.append(section_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # 5. Reliability Metrics
    elements.append(Paragraph('5. RELIABILITY METRICS', styles['Heading2']))
    if reliability_metrics:
        reliability_data = [
            ['Metric', 'Value'],
            ["Cronbach's Alpha", str(reliability_metrics.get('cronbach_alpha', 'N/A'))],
            ['Interpretation', reliability_metrics.get('interpretation', 'N/A')],
        ]
        reliability_table = create_pdf_table(reliability_data)
        elements.append(reliability_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f'analytics_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


def create_pdf_table(data):
    """Create a formatted table for PDF"""
    table = Table(data, hAlign='LEFT')
    table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Data rows
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    return table

