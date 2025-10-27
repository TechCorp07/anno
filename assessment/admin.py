"""
Enhanced Admin Configuration for Assessment App
Includes bulk Excel import, cohort management, and analytics
"""

from django.utils import timezone
from django.contrib import admin
from django.shortcuts import render, redirect
from django.urls import path, reverse
from django.contrib import messages
from django.db.models import Avg, Count, Q
from django.utils.html import format_html
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from pathlib import Path
from django.db import transaction
from django.http import HttpResponse
import openpyxl
from io import BytesIO
import json
import zipfile
import tempfile
import re
import os

from .models import (
    TestCategory, QuestionTopic, Question, Test, TestAttempt, Answer, UserProfile,
    Cohort, CohortMembership, ProctoringEvent, PlagiarismFlag, TestTopicDistribution
)

@admin.register(UserProfile)
class UserProfileAdmin(admin.ModelAdmin):
    list_display = [
        'user', 'phone_number', 'province', 'city', 
        'employment_status', 'education_level', 
        'profile_completed', 'created_at'
    ]
    list_filter = [
        'province', 'employment_status', 'education_level', 
        'has_mri_experience', 'profile_completed'
    ]
    search_fields = [
        'user__username', 'user__email', 'user__first_name', 
        'user__last_name', 'phone_number', 'national_id', 'city'
    ]
    readonly_fields = ['created_at', 'updated_at', 'profile_completed']
    
    fieldsets = (
        ('User Account', {
            'fields': ('user',)
        }),
        ('Personal Information', {
            'fields': (
                'phone_number', 'date_of_birth', 'national_id', 'gender'
            )
        }),
        ('Location', {
            'fields': ('province', 'city', 'address')
        }),
        ('Professional Background', {
            'fields': (
                'employment_status', 'current_employer', 
                'years_of_experience', 'has_mri_experience',
                'education_level', 'institution_attended', 
                'graduation_year'
            )
        }),
        ('Certifications', {
            'fields': ('radiography_license_number', 'license_expiry_date')
        }),
        ('Consents', {
            'fields': (
                'terms_accepted', 'terms_accepted_at',
                'data_processing_consent', 'data_consent_at'
            )
        }),
        ('Profile Status', {
            'fields': ('profile_completed', 'profile_photo')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at')
        }),
    )

class QuestionTopicInline(admin.TabularInline):
    model = QuestionTopic
    extra = 1
    fields = ['name', 'description', 'questions_per_test']


@admin.register(TestCategory)
class TestCategoryAdmin(admin.ModelAdmin):
    list_display = ['name', 'stage_number', 'passing_score', 'is_active', 'question_count']
    list_filter = ['is_active', 'stage_number']
    search_fields = ['name', 'description']
    list_editable = ['is_active']
    inlines = [QuestionTopicInline]
    
    def question_count(self, obj):
        count = sum(topic.questions.count() for topic in obj.topics.all())
        return format_html('<b>{}</b>', count)
    question_count.short_description = 'Total Questions'
    
    fieldsets = (
        ('Category Information', {
            'fields': ('name', 'description', 'stage_number')
        }),
        ('Configuration', {
            'fields': ('passing_score', 'is_active')
        }),
    )


@admin.register(QuestionTopic)
class QuestionTopicAdmin(admin.ModelAdmin):
    list_display = ['name', 'category', 'questions_per_test', 'question_count', 'created_at']
    list_filter = ['category']
    search_fields = ['name', 'description']
    
    def question_count(self, obj):
        return obj.questions.filter(is_active=True).count()
    question_count.short_description = 'Active Questions'


@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    list_display = ['question_text_short', 'topic', 'question_type', 'difficulty_level', 
                    'points', 'time_limit_seconds', 'is_active']
    list_filter = ['question_type', 'difficulty_level', 'is_active', 'topic__category']
    search_fields = ['question_text', 'option_a', 'option_b', 'option_c', 'option_d']
    list_editable = ['is_active', 'difficulty_level']
    change_form_template = 'admin/assessment/question/change_form.html'
    add_form_template = 'admin/assessment/question/change_form.html'
    
    def question_text_short(self, obj):
        text = obj.question_text[:60] + '...' if len(obj.question_text) > 60 else obj.question_text
        
        if obj.question_type in ['dicom', 'image'] and obj.question_image:
            coord_url = reverse('admin:question_set_coordinates', args=[obj.id])
            return format_html(
                '{} <a href="{}" style="margin-left: 10px;" class="button">📍 Set Coordinates</a>',
                text, coord_url
            )
        return text
        
    question_text_short.short_description = 'Question'
    
    fieldsets = (
        ('Question Details', {
            'fields': ('topic', 'question_type', 'question_text', 'question_image', 
                      'difficulty_level', 'time_limit_seconds', 'points')
        }),
        ('Answer Options (MCQ)', {
            'fields': ('option_a', 'option_b', 'option_c', 'option_d', 'correct_answer'),
            'classes': ('collapse',)
        }),
        ('DICOM Question Settings', {
            'fields': ('dicom_file', 'hotspot_coordinates'),
            'classes': ('collapse',)
        }),
        ('Annotation Question Settings', {
            'fields': ('ground_truth_file', 'dice_threshold'),
            'classes': ('collapse',)
        }),
        ('Additional Info', {
            'fields': ('explanation', 'is_active')
        }),
    )
    
    # Bulk import functionality
    change_list_template = 'admin/question_changelist.html'
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('bulk-import/', self.admin_site.admin_view(self.bulk_import_view), 
                 name='question_bulk_import'),
            path('download-template/', self.admin_site.admin_view(self.download_template_view),
                 name='question_download_template'),
            path('download-sample-zip/', self.admin_site.admin_view(self.download_sample_zip_view),
                 name='question_download_sample_zip'),
            path('<int:question_id>/set-coordinates/', self.admin_site.admin_view(self.set_coordinates_view), 
                 name='question_set_coordinates'),
        ]
        return custom_urls + urls
    
    def set_coordinates_view(self, request, question_id):
        """Interactive coordinate picker for DICOM/image questions"""
        question = self.get_object(request, question_id)
        
        if request.method == 'POST':
            coordinates_json = request.POST.get('hotspot_coordinates', '[]')
            try:
                coordinates = json.loads(coordinates_json)
                question.hotspot_coordinates = coordinates
                question.save()
                messages.success(request, f'Hotspot coordinates saved for question #{question_id}')
                return redirect('admin:assessment_question_change', question_id)
            except json.JSONDecodeError:
                messages.error(request, 'Invalid coordinate data')
        
        # Get image URL
        image_url = question.question_image.url if question.question_image else None
        
        if not image_url:
            messages.error(request, 'No image uploaded for this question')
            return redirect('admin:assessment_question_change', question_id)
        
        # Convert existing coordinates to JSON
        existing_hotspots = json.dumps(question.hotspot_coordinates or [])
        
        context = {
            'question': question,
            'image_url': image_url,
            'existing_hotspots': existing_hotspots,
            'opts': self.model._meta,
            'has_view_permission': self.has_view_permission(request, question),
        }
        
        return render(request, 'admin/question_coordinate_picker.html', context)

    def bulk_import_view(self, request):
        """Handle bulk Excel import with image folder"""
        
        if request.method == 'POST':
            excel_file = request.FILES.get('excel_file')
            image_folder = request.FILES.getlist('image_folder')
            zip_file = request.FILES.get('zip_file')
            
            if zip_file:
                return self._process_zip_import(request, zip_file)
            elif excel_file:
                return self._process_standard_import(request, excel_file, image_folder)
            else:
                messages.error(request, 'Please upload either a ZIP file or an Excel file.')
                return redirect('..')
        
        # GET request - show form with stats
        context = {
            'total_questions': Question.objects.count(),
            'total_topics': QuestionTopic.objects.count(),
            'opts': self.model._meta,
            'has_view_permission': self.has_view_permission(request),
        }
        return render(request, 'admin/bulk_import_form.html', context)
    
    def _extract_question_number(self, filename):
        """
        Extract question number from various filename patterns:
        - Q001.png -> 1
        - question_1.png -> 1
        - 001.jpg -> 1
        - q1_image.png -> 1
        """
        patterns = [
            r'[Qq](\d+)',      # Q001, q001
            r'question_?(\d+)',  # question_1, question1
            r'^(\d+)',          # 001.png (starts with number)
            r'_(\d+)[\._]',     # any_1.png or any_1_
        ]
        
        for pattern in patterns:
            match = re.search(pattern, filename)
            if match:
                return int(match.group(1))
        
        return None


    def _process_zip_import(self, request, zip_file):
        """Process ZIP file containing Excel + images"""
        import zipfile
        import tempfile
        
        try:
            # Create temporary directory
            with tempfile.TemporaryDirectory() as temp_dir:
                temp_dir_path = Path(temp_dir)
                
                # Extract ZIP
                with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                    zip_ref.extractall(temp_dir_path)
                
                # Find Excel file
                excel_files = list(temp_dir_path.glob('*.xlsx')) + list(temp_dir_path.glob('*.xls'))
                if not excel_files:
                    messages.error(request, 'No Excel file found in ZIP archive.')
                    return redirect('..')
                
                excel_path = excel_files[0]
                
                # Find image files
                image_extensions = ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.dcm']
                image_files = {}
                
                for ext in image_extensions:
                    for img_path in temp_dir_path.rglob(f'*{ext}'):
                        if not img_path.name.startswith('.'):  # Skip hidden files
                            q_num = self._extract_question_number(img_path.name)
                            if q_num:
                                with open(img_path, 'rb') as f:
                                    image_files[q_num] = (img_path.name, f.read())
                
                # Process with extracted files
                return self._import_questions(
                    request, 
                    str(excel_path), 
                    image_files,
                    is_zip=True
                )
                
        except zipfile.BadZipFile:
            messages.error(request, 'Invalid ZIP file.')
            return redirect('..')
        except Exception as e:
            messages.error(request, f'Error processing ZIP file: {str(e)}')
            return redirect('..')


    def _process_standard_import(self, request, excel_file, image_folder):
        """Process standard upload: separate Excel + images"""
        try:
            # Process images into dictionary
            images_dict = {}
            for img_file in image_folder:
                q_num = self._extract_question_number(img_file.name)
                if q_num:
                    images_dict[q_num] = (img_file.name, img_file.read())
                else:
                    messages.warning(
                        request,
                        f'Could not extract question number from filename: {img_file.name}. Use format Q001.png or 001.png'
                    )
            
            return self._import_questions(request, excel_file, images_dict, is_zip=False)
            
        except Exception as e:
            messages.error(request, f'Error processing files: {str(e)}')
            return redirect('..')


    def _import_questions(self, request, excel_source, images_dict, is_zip=False):
        """Core import logic - works with both ZIP and standard uploads"""
        from django.db import transaction
        from django.core.files.base import ContentFile
        
        try:
            # Read Excel file
            if is_zip:
                wb = openpyxl.load_workbook(excel_source)
            else:
                wb = openpyxl.load_workbook(excel_source)
            
            sheet = wb.active
            
            created_count = 0
            updated_count = 0
            skipped_count = 0
            error_count = 0
            errors = []
            questions_with_images = 0
            
            # Use transaction for database integrity
            with transaction.atomic():
                # Skip header row
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
                    try:
                        # Skip completely empty rows
                        if all(cell.value is None for cell in row):
                            continue
                        
                        # Extract values
                        question_number = int(row[0].value) if row[0].value else None
                        topic_name = str(row[1].value).strip() if row[1].value else None
                        question_type = str(row[2].value).lower().strip() if row[2].value else 'mcq'
                        question_text = str(row[3].value).strip() if row[3].value else None
                        option_a = str(row[4].value).strip() if row[4].value else ''
                        option_b = str(row[5].value).strip() if row[5].value else ''
                        option_c = str(row[6].value).strip() if row[6].value else ''
                        option_d = str(row[7].value).strip() if row[7].value else ''
                        correct_answer = str(row[8].value).lower().strip() if row[8].value else None
                        explanation = str(row[9].value).strip() if row[9].value else ''
                        difficulty = int(row[10].value) if row[10].value else 1
                        time_limit = int(row[11].value) if row[11].value else 60
                        points = int(row[12].value) if row[12].value else 1
                        
                        # Validation
                        if not question_text:
                            errors.append(f'Row {row_idx}: Missing question text')
                            error_count += 1
                            continue
                        
                        if not topic_name:
                            errors.append(f'Row {row_idx}: Missing topic name')
                            error_count += 1
                            continue
                        
                        # Find or create topic
                        try:
                            default_category = TestCategory.objects.filter(is_active=True).first()
                            
                            if not default_category:
                                errors.append(f'Row {row_idx}: No active TestCategory found.')
                                error_count += 1
                                continue
                            
                            topic, created = QuestionTopic.objects.get_or_create(
                                name=topic_name,
                                defaults={
                                    'category': default_category,
                                    'description': 'Auto-created during import',
                                    'questions_per_test': 10
                                }
                            )
                            
                        except Exception as e:
                            errors.append(f'Row {row_idx}: Error with topic: {str(e)}')
                            error_count += 1
                            continue
                        
                        # Validate question type
                        valid_types = ['mcq', 'spatial', 'verbal', 'numerical', 'pattern', 
                                    'error_detection', 'image', 'dicom', 'annotation']
                        if question_type not in valid_types:
                            errors.append(f'Row {row_idx}: Invalid question type "{question_type}"')
                            error_count += 1
                            continue
                        
                        # Validate MCQ fields
                        if question_type == 'mcq':
                            if not all([option_a, option_b, option_c, option_d]):
                                errors.append(f'Row {row_idx}: MCQ questions require all 4 options')
                                error_count += 1
                                continue
                            if correct_answer not in ['a', 'b', 'c', 'd']:
                                errors.append(f'Row {row_idx}: Correct answer must be a, b, c, or d')
                                error_count += 1
                                continue
                        
                        # Create or update question
                        question_data = {
                            'topic': topic,
                            'question_type': question_type,
                            'question_text': question_text,
                            'option_a': option_a,
                            'option_b': option_b,
                            'option_c': option_c,
                            'option_d': option_d,
                            'correct_answer': correct_answer,
                            'explanation': explanation,
                            'difficulty_level': min(max(difficulty, 1), 5),  # Clamp 1-5
                            'time_limit_seconds': max(time_limit, 1),
                            'points': max(points, 1),
                            'is_active': True,
                        }
                        
                        # Check if question already exists (by question_text to avoid duplicates)
                        existing_question = Question.objects.filter(
                            question_text=question_text,
                            topic=topic
                        ).first()
                        
                        if existing_question:
                            # Update existing question
                            for key, value in question_data.items():
                                setattr(existing_question, key, value)
                            question = existing_question
                            question.save()
                            updated_count += 1
                        else:
                            # Create new question
                            question = Question.objects.create(**question_data)
                            created_count += 1
                        
                        # Attach image if available
                        if question_number and question_number in images_dict:
                            filename, image_data = images_dict[question_number]
                            
                            # Determine if it's DICOM or regular image
                            if filename.lower().endswith('.dcm'):
                                question.dicom_file.save(
                                    f'question_{question.id}_{filename}',
                                    ContentFile(image_data),
                                    save=True
                                )
                            else:
                                question.question_image.save(
                                    f'question_{question.id}_{filename}',
                                    ContentFile(image_data),
                                    save=True
                                )
                            
                            questions_with_images += 1
                    
                    except Exception as e:
                        errors.append(f'Row {row_idx}: {str(e)}')
                        error_count += 1
                        continue
            
            # Display results
            success_msg = f'Import completed! Created: {created_count}, Updated: {updated_count}'
            if questions_with_images > 0:
                success_msg += f', With Images: {questions_with_images}'
            if skipped_count > 0:
                success_msg += f', Skipped: {skipped_count}'
            
            messages.success(request, success_msg)
            
            if error_count > 0:
                error_msg = f'{error_count} errors occurred:\n' + '\n'.join(errors[:20])
                if len(errors) > 20:
                    error_msg += f'\n... and {len(errors) - 20} more errors.'
                messages.warning(request, error_msg)
            
            # Show warning for images without questions
            unused_images = set(images_dict.keys()) - set(range(1, sheet.max_row))
            if unused_images:
                messages.info(
                    request,
                    f'{len(unused_images)} image(s) did not match any question numbers: {sorted(list(unused_images)[:10])}'
                )
            
            return redirect('..')
            
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            import traceback
            messages.error(request, traceback.format_exc())
            return redirect('..')


    # Also add this helper view for ZIP creation
    def download_sample_zip_view(self, request):
        """Generate and download sample ZIP with Excel + images"""
        import zipfile
        from io import BytesIO
        import PIL.Image as PILImage
        
        # Create in-memory ZIP
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Add Excel template
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = 'Questions'
            
            # Headers
            headers = [
                'Question Number', 'Topic Name', 'Question Type', 'Question Text',
                'Option A', 'Option B', 'Option C', 'Option D', 
                'Correct Answer (a/b/c/d)', 'Explanation', 
                'Difficulty (1-5)', 'Time Limit (seconds)', 'Points'
            ]
            sheet.append(headers)
            
            # Sample questions
            sheet.append([
                1, 'Verbal Reasoning', 'mcq', 'What is the capital of Zimbabwe?',
                'Harare', 'Bulawayo', 'Mutare', 'Gweru',
                'a', 'Harare is the capital and largest city.',
                1, 60, 1
            ])
            
            sheet.append([
                2, 'Spatial Reasoning', 'image', 'Identify the anatomical structure marked in red.',
                'Liver', 'Kidney', 'Spleen', 'Pancreas',
                'a', 'The red marker indicates the liver.',
                2, 90, 2
            ])
            
            # Style and save Excel to buffer
            from openpyxl.styles import Font, PatternFill
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            zip_file.writestr('questions_template.xlsx', excel_buffer.getvalue())
            
            # Add sample images
            for i in [1, 2]:
                img = PILImage.new('RGB', (400, 300), color=(200, 200, 200))
                img_buffer = BytesIO()
                img.save(img_buffer, format='PNG')
                img_buffer.seek(0)
                zip_file.writestr(f'images/Q{i:03d}.png', img_buffer.getvalue())
            
            # Add README
            readme = """BULK QUESTION IMPORT - README

    FILES IN THIS ZIP:
    - questions_template.xlsx: Fill this with your questions
    - images/: Folder containing sample images

    IMAGE NAMING:
    Images should be named to match the Question Number in Excel:
    - Q001.png matches Question Number 1
    - Q002.png matches Question Number 2
    - Or: 001.png, 002.png (without Q prefix)
    - Or: question_1.png, question_2.png

    SUPPORTED FORMATS:
    - Images: .png, .jpg, .jpeg, .gif, .bmp
    - DICOM: .dcm files

    INSTRUCTIONS:
    1. Fill the Excel template with your questions
    2. Add your images to the images/ folder (or same directory as Excel)
    3. Zip everything together
    4. Upload the ZIP file via the bulk import page

    That's it! The system will:
    - Extract the ZIP
    - Match images to questions by number
    - Import everything in one go
    """
            zip_file.writestr('README.txt', readme)
        
        zip_buffer.seek(0)
        
        response = HttpResponse(
            zip_buffer.getvalue(),
            content_type='application/zip'
        )
        response['Content-Disposition'] = 'attachment; filename=question_import_sample.zip'
        
        return response


    def download_template_view(self, request):
        """Generate and download Excel template"""
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Questions Template'
        
        # Headers
        headers = [
            'Question Number', 'Topic Name', 'Question Type', 'Question Text',
            'Option A', 'Option B', 'Option C', 'Option D', 
            'Correct Answer (a/b/c/d)', 'Explanation', 
            'Difficulty (1-5)', 'Time Limit (seconds)', 'Points'
        ]
        sheet.append(headers)
        
        # Sample data
        sheet.append([
            1, 'Verbal Reasoning', 'mcq', 'What is the capital of Zimbabwe?',
            'Harare', 'Bulawayo', 'Mutare', 'Gweru',
            'a', 'Harare is the capital and largest city of Zimbabwe.',
            1, 60, 1
        ])
        
        # Style header
        from openpyxl.styles import Font, PatternFill
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=question_import_template.xlsx'
        
        return response


@admin.register(Cohort)
class CohortAdmin(admin.ModelAdmin):
    list_display = ['name', 'start_date', 'end_date', 'member_count', 'enabled_categories_list', 'is_active', 'cohort_actions']
    list_filter = ['is_active', 'start_date']
    search_fields = ['name', 'description']
    filter_horizontal = ['enabled_categories']
    list_editable = ['is_active']
    change_list_template = 'admin/cohort_changelist.html'
    
    def member_count(self, obj):
        return obj.members.count()
    member_count.short_description = 'Members'
    
    def enabled_categories_list(self, obj):
        categories = obj.enabled_categories.all()
        return ', '.join([f"Stage {c.stage_number}" for c in categories])
    enabled_categories_list.short_description = 'Enabled Stages'

    def get_urls(self):
        """Add custom URL for bulk user assignment"""
        urls = super().get_urls()
        custom_urls = [
            path('<int:cohort_id>/bulk-add-users/', 
                 self.admin_site.admin_view(self.bulk_add_users_view), 
                 name='cohort_bulk_add_users'),
        ]
        return custom_urls + urls
    
    def bulk_add_users_view(self, request, cohort_id):
        """View for bulk adding users to a cohort"""
        from django.contrib.auth.models import User
        from django.db import transaction
        from django.utils import timezone
        
        cohort = self.get_object(request, cohort_id)
        
        if request.method == 'POST':
            user_ids = request.POST.getlist('users')
            
            if not user_ids:
                messages.error(request, 'No users selected.')
                return redirect(reverse('admin:cohort_bulk_add_users', args=[cohort_id]))
            
            # Bulk create memberships
            try:
                with transaction.atomic():
                    added_count = 0
                    skipped_count = 0
                    
                    for user_id in user_ids:
                        user = User.objects.get(id=user_id)
                        
                        # Check if already a member
                        if CohortMembership.objects.filter(user=user, cohort=cohort).exists():
                            skipped_count += 1
                            continue
                        
                        # Create membership
                        CohortMembership.objects.create(
                            user=user,
                            cohort=cohort,
                            joined_at=timezone.now()
                        )
                        added_count += 1
                    
                    success_msg = f'Successfully added {added_count} user(s) to cohort "{cohort.name}".'
                    if skipped_count > 0:
                        success_msg += f' Skipped {skipped_count} user(s) who were already members.'
                    
                    messages.success(request, success_msg)
                    return redirect(reverse('admin:assessment_cohort_change', args=[cohort_id]))
                    
            except Exception as e:
                messages.error(request, f'Error adding users: {str(e)}')
                return redirect(reverse('admin:cohort_bulk_add_users', args=[cohort_id]))
        
        # GET request - show user selection form
        # Get all users not yet in this cohort
        existing_member_ids = cohort.members.values_list('user_id', flat=True)
        available_users = User.objects.exclude(id__in=existing_member_ids).order_by('username')
        
        # Get current members for display
        current_members = cohort.members.select_related('user').order_by('-joined_at')
        
        context = {
            'cohort': cohort,
            'available_users': available_users,
            'current_members': current_members,
            'title': f'Bulk Add Users to {cohort.name}',
            'opts': self.model._meta,
            'has_change_permission': self.has_change_permission(request, cohort),
        }
        
        return render(request, 'admin/cohort_bulk_add_users.html', context)
    
    def cohort_actions(self, obj):
        """Display action links for cohort"""
        bulk_add_url = reverse('admin:cohort_bulk_add_users', args=[obj.id])
        return format_html(
            '<a href="{}" class="button" style="background-color: #10b981; color: white; padding: 5px 10px; border-radius: 4px;">👥 Bulk Add Users</a>',
            bulk_add_url
        )
    cohort_actions.short_description = 'Actions'
    
@admin.register(CohortMembership)
class CohortMembershipAdmin(admin.ModelAdmin):
    list_display = ['user', 'cohort', 'joined_at']
    list_filter = ['cohort', 'joined_at']
    search_fields = ['user__username', 'user__email', 'user__first_name', 'user__last_name', 'cohort__name']
    raw_id_fields = ['user']
    
    actions = ['bulk_remove_from_cohort']
    list_per_page = 100
    
    def bulk_remove_from_cohort(self, request, queryset):
        """Remove selected users from their cohorts"""
        count = queryset.count()
        queryset.delete()
        self.message_user(
            request,
            f'Successfully removed {count} user(s) from their cohort(s).',
            messages.SUCCESS
        )
    bulk_remove_from_cohort.short_description = 'Remove selected users from cohorts'


class TestTopicDistributionInline(admin.TabularInline):
    """
    Inline admin for configuring question distribution across topics.
    """
    model = TestTopicDistribution
    extra = 1
    fields = ['topic', 'num_questions', 'order']
    verbose_name = 'Topic Distribution'
    verbose_name_plural = 'Question Distribution by Topic'
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """Show all active topics from all categories"""
        if db_field.name == "topic":
            kwargs["queryset"] = QuestionTopic.objects.filter(
                category__is_active=True
            ).select_related('category').order_by('category__name', 'name')
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


@admin.register(Test)
class TestAdmin(admin.ModelAdmin):
    list_display = ['title', 'category', 'time_limit_minutes', 'passing_score', 
                    'total_questions', 'attempt_count', 'is_active']
    list_filter = ['is_active', 'category', 'require_webcam', 'require_fullscreen']
    search_fields = ['title', 'description']
    filter_horizontal = ['manual_questions']
    list_editable = ['is_active']
    change_form_template = 'admin/assessment/test/change_form.html'
    add_form_template = 'admin/assessment/test/change_form.html'
    inlines = [TestTopicDistributionInline]
    
    def total_questions(self, obj):
        return obj.get_total_questions()
    total_questions.short_description = 'Questions'
    
    def attempt_count(self, obj):
        return obj.attempts.count()
    attempt_count.short_description = 'Attempts'
    
    fieldsets = (
        ('Test Information', {
            'fields': ('category', 'title', 'description')
        }),
        ('Question Configuration', {
            'fields': ('auto_generate_from_topics', 'manual_questions')
        }),
        ('Test Configuration', {
            'fields': ('time_limit_minutes', 'passing_score', 'is_active')
        }),
        ('Proctoring Settings', {
            'fields': ('require_webcam', 'require_fullscreen', 'snapshot_interval_seconds')
        }),
    )


class AnswerInline(admin.TabularInline):
    model = Answer
    extra = 0
    readonly_fields = ['question', 'selected_answer', 'is_correct', 'time_spent_seconds', 'answered_at']
    can_delete = False
    
    def has_add_permission(self, request, obj=None):
        return False


@admin.register(TestAttempt)
class TestAttemptAdmin(admin.ModelAdmin):
    list_display = ['user', 'test', 'cohort', 'status', 'score_display', 'passed', 
                    'flagged_for_plagiarism', 'started_at', 'consent_status', 'view_proctoring']
    list_filter = ['status', 'passed', 'flagged_for_plagiarism','consent_given', 'test__category', 'started_at']
    search_fields = ['user__username', 'test__title', 'ip_address']
    readonly_fields = ['user', 'test', 'started_at', 'completed_at', 'time_spent_seconds', 
                       'ip_address', 'user_agent', 'similarity_score']
    inlines = [AnswerInline]
    
    def consent_status(self, obj):
        """Display consent acceptance status"""
        if obj.consent_given:
            return format_html(
                '<span style="color: #10b981;">✓ Accepted</span><br><small>{}</small>',
                obj.consent_timestamp.strftime('%Y-%m-%d %H:%M') if obj.consent_timestamp else ''
            )
        return format_html('<span style="color: #dc2626;">✗ Not Given</span>')
    consent_status.short_description = 'Consent'
    
    def view_proctoring(self, obj):
        """Link to view all proctoring events and images"""
        count = obj.proctoring_events.count()
        critical = obj.proctoring_events.filter(severity='critical').count()
        images_count = obj.proctoring_events.filter(image_file__isnull=False).count()
        
        # Link to events list
        events_url = f'/admin/assessment/proctoringevent/?attempt__id__exact={obj.id}'
        # Link to image gallery
        gallery_url = f'/proctoring/images/{obj.id}/'
        
        if critical > 0:
            return format_html(
                '<a href="{}" style="color: #dc2626; font-weight: bold;">🚨 {} events ({} critical)</a><br>'
                '<a href="{}" style="color: #3b82f6; font-weight: bold;">📸 View {} Images</a>',
                events_url, count, critical, gallery_url, images_count
            )
        return format_html(
            '<a href="{}">{} events</a><br>'
            '<a href="{}" style="color: #3b82f6;">📸 View {} Images</a>',
            events_url, count, gallery_url, images_count
        )
    view_proctoring.short_description = 'Proctoring'
    
    def status_display(self, obj):
        """Display status with color coding"""
        colors = {
            'completed': '#10b981',
            'in_progress': '#3b82f6',
            'flagged': '#dc2626',
            'expired': '#6b7280',
        }
        return format_html(
            '<span style="color: {}; font-weight: 600;">{}</span>',
            colors.get(obj.status, '#000'),
            obj.get_status_display()
        )
    status_display.short_description = 'Status'
    
    fieldsets = (
        ('Attempt Details', {
            'fields': ('user', 'test', 'cohort', 'status')
        }),
        ('Results', {
            'fields': ('score', 'passed')
        }),
        ('Proctoring & Consent', {
            'fields': (
                'consent_given', 
                'consent_timestamp',
                'ip_address', 
                'user_agent',
            ),
            'description': 'Consent acceptance and proctoring information'
        }),
        ('Plagiarism Detection', {
            'fields': ('similarity_score', 'flagged_for_plagiarism')
        }),
        ('Timing', {
            'fields': ('started_at', 'completed_at', 'time_spent_seconds')
        }),
    )
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('analytics/', 
                 self.admin_site.admin_view(self.analytics_dashboard_view),
                 name='testattempt_analytics'),
            path('export-analytics/', 
                 self.admin_site.admin_view(self.export_analytics_view),
                 name='testattempt_export'),
        ]
        return custom_urls + urls
    
    def analytics_dashboard_view(self, request):
        """Full analytics dashboard"""
        from assessment.analytics_views import admin_analytics_dashboard
        return admin_analytics_dashboard(request)
    
    def export_analytics_view(self, request):
        """Export analytics to Excel"""
        from assessment.analytics_views import export_analytics_excel
        return export_analytics_excel(request)
    
    def score_display(self, obj):
        if obj.score is not None:
            color = 'green' if obj.passed else 'red'
            score_str = f'{float(obj.score):.2f}%'
            return format_html('<b style="color: {};">{}</b>', color, score_str)
        return '-'
    
    score_display.short_description = 'Score'
    
    fieldsets = (
        ('Attempt Details', {
            'fields': ('user', 'test', 'cohort', 'status')
        }),
        ('Results', {
            'fields': ('score', 'passed')
        }),
        ('Proctoring', {
            'fields': ('ip_address', 'user_agent', 'consent_given', 'consent_timestamp')
        }),
        ('Plagiarism Detection', {
            'fields': ('similarity_score', 'flagged_for_plagiarism')
        }),
        ('Timing', {
            'fields': ('started_at', 'completed_at', 'time_spent_seconds')
        }),
    )
    
    # Analytics dashboard
    change_list_template = 'admin/test_attempt_changelist.html'
    
    def changelist_view(self, request, extra_context=None):
        """Add analytics to changelist"""
        extra_context = extra_context or {}
        
        # Calculate statistics
        qs = self.get_queryset(request)
        completed = qs.filter(status='completed')
        
        extra_context['total_attempts'] = qs.count()
        extra_context['completed_attempts'] = completed.count()
        extra_context['pass_rate'] = completed.filter(passed=True).count() / completed.count() * 100 if completed.count() > 0 else 0
        
        # Convert Decimal to float for proper template formatting
        avg_score_decimal = completed.aggregate(Avg('score'))['score__avg']
        extra_context['avg_score'] = float(avg_score_decimal) if avg_score_decimal is not None else 0
        
        extra_context['flagged_count'] = qs.filter(flagged_for_plagiarism=True).count()
        
        # Per-category statistics
        category_stats = []
        for category in TestCategory.objects.all():
            cat_attempts = completed.filter(test__category=category)
            if cat_attempts.exists():
                avg_score_decimal = cat_attempts.aggregate(Avg('score'))['score__avg']
                category_stats.append({
                    'name': category.name,
                    'stage': category.stage_number,
                    'attempts': cat_attempts.count(),
                    'pass_rate': cat_attempts.filter(passed=True).count() / cat_attempts.count() * 100,
                    'avg_score': float(avg_score_decimal) if avg_score_decimal is not None else 0,
                })
        extra_context['category_stats'] = category_stats
        
        return super().changelist_view(request, extra_context=extra_context)

@admin.register(ProctoringEvent)
class ProctoringEventAdmin(admin.ModelAdmin):
    """
    Enhanced admin for proctoring events with severity highlighting
    """
    list_display = [
        'severity_icon',
        'attempt_user',
        'event_type_display', 
        'severity',
        'timestamp',
        'has_image',
        'view_details'
    ]
    list_filter = [
        'severity',
        'event_type',
        'timestamp',
        'attempt__test',
    ]
    search_fields = [
        'attempt__user__username',
        'attempt__user__email',
        'event_type',
        'metadata',
    ]
    readonly_fields = [
        'attempt',
        'event_type',
        'image_file',
        'metadata',
        'severity',
        'timestamp',
        'formatted_metadata',
        'image_preview',
    ]
    date_hierarchy = 'timestamp'
    
    def severity_icon(self, obj):
        """Display colored icon based on severity"""
        icons = {
            'critical': '🚨',
            'warning': '⚠️',
            'info': 'ℹ️',
        }
        colors = {
            'critical': '#dc2626',
            'warning': '#f59e0b',
            'info': '#3b82f6',
        }
        return format_html(
            '<span style="font-size: 20px; color: {};">{}</span>',
            colors.get(obj.severity, '#000'),
            icons.get(obj.severity, '•')
        )
    severity_icon.short_description = ''
    
    def attempt_user(self, obj):
        """Display username with link to test attempt"""
        url = f'/admin/assessment/testattempt/{obj.attempt.id}/change/'
        return format_html(
            '<a href="{}">{}</a>',
            url,
            obj.attempt.user.username
        )
    attempt_user.short_description = 'User'
    
    def event_type_display(self, obj):
        """Display event type with color coding"""
        colors = {
            'camera_disabled': '#dc2626',
            'camera_permission_denied': '#dc2626',
            'tab_switch': '#f59e0b',
            'fullscreen_exit': '#f59e0b',
            'camera_access_granted': '#10b981',
            'proctoring_initialized': '#10b981',
            'consent_accepted': '#10b981',
        }
        color = colors.get(obj.event_type, '#6b7280')
        return format_html(
            '<span style="color: {}; font-weight: 600;">{}</span>',
            color,
            obj.get_event_type_display()
        )
    event_type_display.short_description = 'Event Type'
    
    def has_image(self, obj):
        """Show if event has image attachment"""
        if obj.image_file:
            return format_html(
                '<span style="color: #10b981;">✓ Image</span>'
            )
        return format_html(
            '<span style="color: #9ca3af;">-</span>'
        )
    has_image.short_description = 'Snapshot'
    
    def view_details(self, obj):
        """Link to view full details"""
        url = f'/admin/assessment/proctoringevent/{obj.id}/change/'
        return format_html(
            '<a href="{}">View Details →</a>',
            url
        )
    view_details.short_description = 'Actions'
    
    def formatted_metadata(self, obj):
        """Display metadata in readable format"""
        import json
        if obj.metadata:
            formatted = json.dumps(obj.metadata, indent=2)
            return format_html('<pre style="background: #f3f4f6; padding: 10px; border-radius: 4px;">{}</pre>', formatted)
        return '-'
    formatted_metadata.short_description = 'Event Metadata'
    
    def image_preview(self, obj):
        """Show image preview if available"""
        if obj.image_file:
            return format_html(
                '<img src="{}" style="max-width: 400px; max-height: 400px; border: 1px solid #ddd; border-radius: 4px;" />',
                obj.image_file.url
            )
        return 'No image'
    image_preview.short_description = 'Image Preview'
    
    fieldsets = (
        ('Event Information', {
            'fields': ('attempt', 'event_type', 'severity', 'timestamp')
        }),
        ('Image Data', {
            'fields': ('image_file', 'image_preview'),
            'classes': ('collapse',),
        }),
        ('Metadata', {
            'fields': ('formatted_metadata',),
            'classes': ('collapse',),
        }),
    )
    
    # Custom action to mark critical events as reviewed
    actions = ['mark_as_reviewed']
    
    def mark_as_reviewed(self, request, queryset):
        # Add a 'reviewed' flag to metadata
        count = 0
        for event in queryset:
            if not event.metadata:
                event.metadata = {}
            event.metadata['reviewed_by'] = request.user.username
            event.metadata['reviewed_at'] = timezone.now().isoformat()
            event.save()
            count += 1
        self.message_user(request, f'{count} events marked as reviewed.')
    mark_as_reviewed.short_description = 'Mark selected events as reviewed'


@admin.register(PlagiarismFlag)
class PlagiarismFlagAdmin(admin.ModelAdmin):
    list_display = ['attempt1_user', 'attempt2_user', 'similarity_percentage', 
                    'reviewed', 'action_taken', 'detected_at']
    list_filter = ['reviewed', 'action_taken', 'detected_at']
    search_fields = ['attempt1__user__username', 'attempt2__user__username']
    readonly_fields = ['attempt1', 'attempt2', 'similarity_percentage', 
                       'matching_answers', 'detected_at']
    
    def attempt1_user(self, obj):
        return obj.attempt1.user.username
    attempt1_user.short_description = 'User 1'
    
    def attempt2_user(self, obj):
        return obj.attempt2.user.username
    attempt2_user.short_description = 'User 2'
    
    fieldsets = (
        ('Plagiarism Detection', {
            'fields': ('attempt1', 'attempt2', 'similarity_percentage', 'matching_answers')
        }),
        ('Review', {
            'fields': ('reviewed', 'reviewer_notes', 'action_taken', 'reviewed_at')
        }),
        ('Timestamps', {
            'fields': ('detected_at',)
        }),
    )

def admin_dashboard_view(request):
    """
    Central admin dashboard - hub for all admin features
    """
    from django.contrib.auth.models import User
    
    # Gather statistics
    total_questions = Question.objects.filter(is_active=True).count()
    total_users = User.objects.count()
    total_attempts = TestAttempt.objects.filter(status='completed').count()
    
    pass_rate = 0
    if total_attempts > 0:
        passed = TestAttempt.objects.filter(passed=True).count()
        pass_rate = (passed / total_attempts) * 100
    
    flagged_count = TestAttempt.objects.filter(flagged_for_plagiarism=True).count()
    
    # Recent activity
    recent_attempts = TestAttempt.objects.select_related('user', 'test').order_by('-started_at')[:50]
    
    context = {
        'total_questions': total_questions,
        'total_users': total_users,
        'total_attempts': total_attempts,
        'pass_rate': pass_rate,
        'flagged_count': flagged_count,
        'recent_attempts': recent_attempts,
    }
    
    return render(request, 'admin/admin_dashboard.html', context)


# Register the custom dashboard URL with Django Admin
_original_get_urls = admin.site.get_urls

def custom_get_urls():
    """Add custom dashboard URL to admin site"""
    from django.urls import path
    
    urls = _original_get_urls()
    custom_urls = [
        path('dashboard/', admin.site.admin_view(admin_dashboard_view), name='admin_dashboard'),
    ]
    return custom_urls + urls

# Override admin site's get_urls method
admin.site.get_urls = custom_get_urls
